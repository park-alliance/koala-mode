"""
One-time import: reads Workout_Tracker_V1.xlsx and writes data-seed.json,
which the app loads into localStorage the first time it runs.
"""
import json
import re
import openpyxl

SRC = r"C:\Users\josep\coding projects\workout-tracker-app\data\Workout_Tracker_V1.xlsx"
OUT = r"C:\Users\josep\coding projects\workout-tracker-app\data-seed.js"

STANDARD_SHEETS = ["Legs", "Arms", "Chest", "Back", "Shoulders", "Core"]
CARDIO_SHEET = "Cardio"

# Known bad date entries, confirmed with Joseph and corrected here.
# Chest row 113-115 ("02/72/26" isn't a real date) -> actually 7/27/26.
DATE_FIXES = {
    ("Chest", "02/72/26"): "2026-07-27",
}

def slugify(text):
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")

wb = openpyxl.load_workbook(SRC, data_only=True)

categories = STANDARD_SHEETS + [CARDIO_SHEET]

# 1. Seed exercise names per category from the "Exercise List" sheet (preserves order)
exercise_names = {cat: [] for cat in categories}
ex_list_ws = wb["Exercise List"]
header = [c.value for c in ex_list_ws[2]]
col_for_cat = {name: idx for idx, name in enumerate(header) if name in categories}
for row in ex_list_ws.iter_rows(min_row=3, values_only=True):
    for cat, idx in col_for_cat.items():
        if idx < len(row) and row[idx]:
            name = str(row[idx]).strip()
            if name not in exercise_names[cat]:
                exercise_names[cat].append(name)

# 2. Any exercise used in a log sheet but missing from the list also gets added
for cat in STANDARD_SHEETS:
    ws = wb[cat]
    for row in ws.iter_rows(min_row=3, values_only=True):
        name = row[1]
        if name and str(name).strip() not in exercise_names[cat]:
            exercise_names[cat].append(str(name).strip())

ws = wb[CARDIO_SHEET]
for row in ws.iter_rows(min_row=3, values_only=True):
    name = row[1]
    if name and str(name).strip() not in exercise_names[CARDIO_SHEET]:
        exercise_names[CARDIO_SHEET].append(str(name).strip())

exercises = []
exercise_id_by_cat_name = {}
for cat in categories:
    for name in exercise_names[cat]:
        ex_id = f"{slugify(cat)}__{slugify(name)}"
        exercises.append({"id": ex_id, "name": name, "category": cat})
        exercise_id_by_cat_name[(cat, name)] = ex_id

# 3. Logs: standard strength sheets (Date, Exercise, Set, Weight, Reps, Comment)
logs = []
skipped = []
log_id = 1
for cat in STANDARD_SHEETS:
    ws = wb[cat]
    for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        date, ex_name, set_num, weight, reps, comment = row
        if date is None or ex_name is None:
            continue
        if not hasattr(date, "isoformat"):
            fixed = DATE_FIXES.get((cat, str(date)))
            if fixed is None:
                skipped.append({"sheet": cat, "row": row_num, "reason": f"invalid date value: {date!r}"})
                continue
            date_str = fixed
        else:
            date_str = date.date().isoformat()
        ex_id = exercise_id_by_cat_name.get((cat, str(ex_name).strip()))
        if ex_id is None:
            skipped.append({"sheet": cat, "row": row_num, "reason": f"unmatched exercise: {ex_name!r}"})
            continue
        logs.append({
            "id": log_id,
            "date": date_str,
            "exerciseId": ex_id,
            "set": int(set_num) if set_num is not None else None,
            "weight": float(weight) if weight is not None else None,
            "reps": float(reps) if reps is not None else None,
            "comment": comment,
        })
        log_id += 1

# 4. Cardio sheet has different columns: Date, Exercise, Type, Total Time, Work, Work Zone, Rest, Rest Zone, Rounds, Comment
ws = wb[CARDIO_SHEET]
for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
    date, ex_name = row[0], row[1]
    if date is None or ex_name is None:
        continue
    if not hasattr(date, "isoformat"):
        skipped.append({"sheet": CARDIO_SHEET, "row": row_num, "reason": f"invalid date value: {date!r}"})
        continue
    ex_id = exercise_id_by_cat_name.get((CARDIO_SHEET, str(ex_name).strip()))
    if ex_id is None:
        skipped.append({"sheet": CARDIO_SHEET, "row": row_num, "reason": f"unmatched exercise: {ex_name!r}"})
        continue
    _, _, ctype, total_time, work, work_zone, rest, rest_zone, rounds, comment = row
    is_interval = ctype and "interval" in str(ctype).lower()

    segments = []
    if is_interval:
        rounds_n = int(rounds) if rounds else 1
        for _ in range(rounds_n):
            if work:
                segments.append({"time": round(float(work) / 60, 2), "zone": work_zone})
            if rest:
                segments.append({"time": round(float(rest) / 60, 2), "zone": rest_zone})
    elif work_zone:
        segments.append({"time": float(total_time) if total_time else None, "zone": work_zone})

    logs.append({
        "id": log_id,
        "date": date.date().isoformat(),
        "exerciseId": ex_id,
        "totalTime": float(total_time) if total_time else None,
        "distance": None,
        "segments": segments,
        "comment": str(comment) if comment else None,
    })
    log_id += 1

seed = {
    "categories": categories,
    "exercises": exercises,
    "logs": logs,
}

with open(OUT, "w", encoding="utf-8") as f:
    f.write("// Auto-generated by scripts/extract_seed.py from data/Workout_Tracker_V1.xlsx\n")
    f.write("// Used only the first time the app runs, to pre-fill localStorage.\n")
    f.write("const SEED_DATA = ")
    json.dump(seed, f, indent=2)
    f.write(";\n")

print(f"categories: {len(categories)}")
print(f"exercises: {len(exercises)}")
print(f"logs: {len(logs)}")
print(f"skipped: {len(skipped)}")
for s in skipped:
    print("  SKIPPED:", s)
